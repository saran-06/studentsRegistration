from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from tkinter.ttk import Combobox

background_all="#06283D"
framebg="#EDEDED"
framefg="#06283D"

main=Tk()
main.title("Student Registration System")
main.geometry("1250x700+210+100")
main.config(bg=background_all)
file=pathlib.Path("student_deatils.xlsx")
if file.exists():
   	pass
else:
	file=Workbook()
	sheet=file.active
	sheet['A1']="Registration number"
	sheet['B1']="Name"
	sheet['C1']="Department"
	sheet['D1']="Gender"
	sheet['E1']="DOB"
	sheet['F1']="Date of Registration"
	sheet['G1']="Religion"
	sheet['H1']="Cut off"
	sheet['I1']="Father Name"
	sheet['J1']="Mother Name"
	sheet['K1']="Father's occupation"
	sheet['L1']="Mother's occupation"
	file.save("student_deatils.xlsx")

#Exit-function
def Exit():
    	main.destroy()
#gender-function
def selection():
	global gender
	value=radio.get()
	if value==1:
	 	gender="Male"
	else:
		gender="Female"
#upload_image
def show_image():
	global filename
	global pro_image1
	filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="select image file",filetype=(("JPG File","*.jpg"),("PNG File","*.png"),("All files","*.txt")))
	pro_image1=(Image.open(filename))
	resize_image=pro_image1.resize((190,190))
	photo_pro=ImageTk.PhotoImage(resize_image)
	pro_image1_label.config(image=photo_pro)
	pro_image1_label.image=photo_pro

#automatic regno updation
def registeration_no():
	file=openpyxl.load_workbook("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
	sheet=file.active
	row=sheet.max_row
	max_row_val=sheet.cell(row=row,column=1).value
	try:
		Registration.set(max_row_val+1)
	except:
		Registration.set("1")
#clear-reset
def clear():
	Name.set("")
	DOB.set("")
	Religion.set("")
	Cut_off.set("")
	F_name.set("")
	F_occ.set("")
	M_name.set("")
	M_occ.set("")
	Department.set("Select department")
	registeration_no()

	Save_but.config(state='normal')
	img1=PhotoImage(file="C:\\Users\\Saranya\\Documents\\student registeration page\\images\\upload photo (1).png")
	pro_image1_label.config(image=img1)
	pro_image1_label.image=img1
	
	pro_image1=""

#save
def save():
	R1=Registration.get()
	N1=Name.get()
	C1=Department.get()
	try:
		G1=gender
	except:
		messagebox.showerror("error","Select Gender!")
	D2=DOB.get()
	D1=Date.get()
	Re1=Religion.get()
	De1=Department.get()
	fathername=F_name.get()
	mothername=M_name.get()
	F1=F_occ.get()
	M1=M_occ.get()

	if N1=="" or C1=="Select department" or D2=="" or Re1=="" or De1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
		messagebox.showerror("Few data is missing!")
	else:
		file=openpyxl.load_workbook("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
		sheet=file.active
		sheet.cell(column=1,row=sheet.max_row+1,value=R1)
		sheet.cell(column=2,row=sheet.max_row,value=N1)
		sheet.cell(column=3,row=sheet.max_row,value=C1)
		sheet.cell(column=4,row=sheet.max_row,value=G1)

		sheet.cell(column=5,row=sheet.max_row,value=D2)
		sheet.cell(column=6,row=sheet.max_row,value=D1)
		sheet.cell(column=7,row=sheet.max_row,value=Re1)
		sheet.cell(column=8,row=sheet.max_row,value=De1)
		sheet.cell(column=9,row=sheet.max_row,value=fathername)
		sheet.cell(column=10,row=sheet.max_row,value=mothername)
		sheet.cell(column=11,row=sheet.max_row,value=F1)
		sheet.cell(column=12,row=sheet.max_row,value=M1)
		file.save("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
		try:
			pro_image1.save("C:\\Users\\Saranya\\Documents\\student registeration page\\student_images\\"+str(R1)+".jpg")
		except:
			messagebox.showinfo("info","Profile picture is not available!")
		messagebox.showinfo("info","Sucessfully data entered!")
		clear()	
		registeration_no()

#search
def search():
	text=Search.get()
	clear()
	Save_but.config(state="disable")
	file=openpyxl.load_workbook("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
	sheet=file.active
	for row in sheet.rows:
		if row[0].value==int(text):
			sheet_name=row[0]
			reg_position=str(sheet_name)[14:-1]
			reg_no=str(sheet_name)[15:-1]
	try:
		print(str(sheet_name))
	except:
		messagebox.showerror("Invalid","Invalid registeration number!")
	x1=sheet.cell(row=int(reg_no),column=1).value
	x2=sheet.cell(row=int(reg_no),column=2).value
	x3=sheet.cell(row=int(reg_no),column=3).value
	x4=sheet.cell(row=int(reg_no),column=4).value
	x5=sheet.cell(row=int(reg_no),column=5).value
	x6=sheet.cell(row=int(reg_no),column=6).value
	x7=sheet.cell(row=int(reg_no),column=7).value
	x8=sheet.cell(row=int(reg_no),column=8).value
	x9=sheet.cell(row=int(reg_no),column=9).value
	x10=sheet.cell(row=int(reg_no),column=10).value
	x11=sheet.cell(row=int(reg_no),column=11).value
	x12=sheet.cell(row=int(reg_no),column=12).value
	print(x4)
	
	Registration.set(x1)
	Name.set(x2)
	Department.set(x3)
	if x4=="male":
		r1.select()
	else:
		r2.select()
	DOB.set(x5)
	Date.set(x6)
	Religion.set(x7)
	Cut_off.set(x8)
	F_name.set(x9)
	M_name.set(x10)
	F_occ.set(x11)
	M_occ.set(x12)
	pro_image1=(Image.open("C:\\Users\\Saranya\\Documents\\student registeration page\\student_images\\"+str(x1)+".jpg"))
	resize_image=pro_image1.resize((190,190))
	photo2=ImageTk.PhotoImage(resize_image)
	pro_image1_label.config(image=photo2)
	pro_image1_label.image=photo2
	
#Upadte fn
def Update():
	R1=Registration.get()
	N1=Name.get()
	C1=Department.get()
	selection()
	G1=gender
	D2=DOB.get()
	D1=Date.get()
	Re1=Religion.get()
	De1=Department.get()
	fathername=F_name.get()
	mothername=M_name.get()
	F1=F_occ.get()
	M1=M_occ.get()
	file=openpyxl.load_workbook("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
	sheet=file.active
	for row in sheet.rows:
		if row[0].value==R1:
			sheet_name=row[0]
			reg_position=str(sheet_name)[14:-1]
			reg_no=str(sheet_name)[15:-1]
	sheet.cell(column=2,row=int(reg_no),value=N1)
	sheet.cell(column=3,row=int(reg_no),value=C1)
	sheet.cell(column=4,row=int(reg_no),value=G1)
	sheet.cell(column=5,row=int(reg_no),value=D2)
	sheet.cell(column=6,row=int(reg_no),value=D1)
	sheet.cell(column=7,row=int(reg_no),value=Re1)
	sheet.cell(column=8,row=int(reg_no),value=De1)
	sheet.cell(column=9,row=int(reg_no),value=fathername)
	sheet.cell(column=10,row=int(reg_no),value=mothername)
	sheet.cell(column=11,row=int(reg_no),value=F1)
	sheet.cell(column=12,row=int(reg_no),value=M1)
	file.save("C:\\Users\\Saranya\\Documents\\student registeration page\\student_deatils.xlsx")
	try:
			pro_image1.save("C:\\Users\\Saranya\\Documents\\student registeration page\\student_images\\"+str(R1)+".jpg")
	except:
		pass
	messagebox.showinfo("Update","Update Successfully!")
	clear()

	

	



#frames
Label(main,text="Email:Bannariamman.bitsathy.ac.in",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(main,text="STUDENT REGISTERATION FORM", width=10,height=4,bg="brown",anchor="center",font="Timesroman 15 bold",fg="white").pack(side=TOP,fill=X)

#search box
Search=StringVar()
Entry(main,textvariable=Search,width=15,bd=4,font="Timesroman").place(x=810,y=87)
image_icon=PhotoImage(file="C:\\Users\\Saranya\\Documents\\student registeration page\\images\\search.png")
srch=Button(main,text="Search",compound=LEFT,image=image_icon,width=140,bg="skyblue",font="Timesroman",command=search)
srch.place(x=1000,y=81)
image_update=PhotoImage(file="C:\\Users\\Saranya\\Documents\\student registeration page\\images\\Layer 4 (1).png")
update_button=Button(main,image=image_update,bg="#c36464",command=Update)
update_button.place(x=140,y=70)

#regno and date
Label(main,text="Registration number:",fg=framebg,bg=background_all).place(x=30,y=180)
Label(main,text="Date:",fg=framebg,bg=background_all).place(x=500,y=180)

Registration=IntVar()
Date=StringVar()

reg_entry=Entry(main,textvariable=Registration,width=15,font="Timesroman 10")
reg_entry.place(x=160,y=180)

registeration_no()

#setting date
today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(main,textvariable=Date,width=15,font="Timesroman 10")
date_entry.place(x=550,y=180)
Date.set(d1)

#student_details
stu_obj=LabelFrame(main,text="Student's Details",width=900,height=220,bg=framebg,fg=framefg,relief=GROOVE)
stu_obj.place(x=30,y=220)

#sublabels of stu_obj
Label(stu_obj,text="Full Name:",font="Timesroman 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(stu_obj,text="Date of Birth:",font="Timesroman 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(stu_obj,text="Gender:",font="Timesroman 13",bg=framebg,fg=framefg).place(x=30,y=150)
Label(stu_obj,text="Department:",font="Timesroman 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(stu_obj,text="Religion:",font="Timesroman 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(stu_obj,text="Cut off",font="Timesroman 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
reg_entry=Entry(stu_obj,textvariable=Name,width=20,bd=4,font="Timesroman 10")
reg_entry.place(x=160,y=50)

DOB=StringVar()
DOB_entry=Entry(stu_obj,textvariable=DOB,width=20,bd=4,font="Timesroman 10")
DOB_entry.place(x=160,y=100)

radio=IntVar()
r1=Radiobutton(stu_obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
r1.place(x=150,y=150)

r2=Radiobutton(stu_obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
r2.place(x=200,y=150)

Religion=StringVar()
Religion_entry=Entry(stu_obj,textvariable=Religion,width=20,bd=4,font="Timesroman 10")
Religion_entry.place(x=630,y=100)

Cut_off=StringVar()
Cut_off_entry=Entry(stu_obj,textvariable=Cut_off,width=20,bd=4,font="Timesroman 10")
Cut_off_entry.place(x=630,y=150)

Department=Combobox(stu_obj,values=["Aero", "AI&DS", "AI&ML", "Agri", "Biotech", "CSE", "CT", "IT", "CSBS", "Food Tech", "Textile", "ISE"],font="arial 10",width=17,state="r")
Department.place(x=630,y=50)
Department.set("Select department")


#parent_details
par_obj=LabelFrame(main,text="Parent's Details",width=900,height=220,bg=framebg,fg=framefg,relief=GROOVE)
par_obj.place(x=30,y=490)

Label(par_obj,text="Father's Name:",font="Timesromen",bg=framebg,fg=framefg).place(x=30,y=50)
Label(par_obj,text="Occupation",font="Timesromen",bg=framebg,fg=framefg).place(x=30,y=100)

F_name=StringVar()
F_name_entry=Entry(par_obj,textvariable=F_name,bd=4,width=20,font="Timesroman")
F_name_entry.place(x=160,y=50)


F_occ=StringVar()
F_occ_entry=Entry(par_obj,textvariable=F_occ,bd=4,width=20,font="Timesroman")
F_occ_entry.place(x=160,y=100)

Label(par_obj,text="Mother's Name:",font="Timesromen",bg=framebg,fg=framefg).place(x=500,y=50)
Label(par_obj,text="Occupation",font="Timesromen",bg=framebg,fg=framefg).place(x=500,y=100)

M_name=StringVar()
M_name_entry=Entry(par_obj,textvariable=M_name,bd=4,width=20,font="Timesroman")
M_name_entry.place(x=630,y=50)


M_occ=StringVar()
M_occ_entry=Entry(par_obj,textvariable=M_occ,bd=4,width=20,font="Timesroman")
M_occ_entry.place(x=630,y=100)

#image
pro_image=Frame(main,bg="black",width=200,height=200,bd=3,relief=GROOVE)
pro_image.place(x=1000,y=150)

pro_image1=PhotoImage(file="C:\\Users\\Saranya\\Documents\\student registeration page\\images\\upload photo (1).png")
pro_image1_label=Label(pro_image,bg="black",image=pro_image1)
pro_image1_label.place(x=0,y=0)

#button_upload
upload_but=Button(main,text="Upload",bg="lightblue",width=19,height=2,font="Timesroman",command=show_image)
upload_but.place(x=1010,y=370)

Save_but=Button(main,text="Save",bg="lightgreen",width=19,height=2,font="Timesroman",command=save)
Save_but.place(x=1010,y=450)

Reset_but=Button(main,text="Reset",bg="lightpink",width=19,height=2,font="Timesroman",command=clear)
Reset_but.place(x=1010,y=530)

Exit_but=Button(main,text="Exit",bg="grey",width=19,height=2,font="Timesroman",command=Exit)
Exit_but.place(x=1010,y=610)

 	


main.mainloop()